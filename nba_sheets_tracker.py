#!/usr/bin/env python3
"""
nba_sheets_tracker.py  -  Track EVERY tab of a Google Sheet and log row-level changes.

Pulls the whole spreadsheet in one request via the /export?format=xlsx endpoint (which contains
all tabs), reads each tab with openpyxl, and diffs its rows against the previous snapshot.
Added/removed rows are appended to nba_sheet_snapshots/sheets_log.csv, which nba_changelog.py
folds into the same report as the players / transactions / docs feeds. Each tab is labelled by
its tab name, so a 30-team workbook produces 30 clearly-named feeds automatically.

First run per workbook saves a baseline silently, like the other trackers.

REQUIRES:  pip install openpyxl

USAGE
-----
    python nba_sheets_tracker.py            # one pass, all tabs
    python nba_sheets_tracker.py --test     # dry run: show per-tab row counts, don't log
    python nba_sheets_tracker.py --silent   # no toast

Exposes poll_once() and send_test() so nba_tracker_all.py drives it with the others.
The spreadsheet must be viewable by "anyone with the link" for the export to work without login.
"""

import os
import re
import io
import csv
import sys
import json
import datetime
import urllib.request

try:
    from winotify import Notification, audio
    HAVE_TOAST = True
except Exception:
    HAVE_TOAST = False

BASE = os.path.dirname(os.path.abspath(__file__))
SNAP = os.path.join(BASE, "nba_sheet_snapshots")
os.makedirs(SNAP, exist_ok=True)
LOG = os.path.join(SNAP, "sheets_log.csv")

# Spreadsheets to watch (ALL tabs of each). Optional friendly prefix; leave "" for none.
SPREADSHEETS = [
    ("", "https://docs.google.com/spreadsheets/d/1w21QbBrNnDyesZnyz13LMzRngInTnCt1/edit?gid=476033577#gid=476033577"),
]
# Tab names to ignore (e.g. a summary/index tab). Case-sensitive.
SKIP_TABS = ["NBA Salary Cap Report", "Aggregate Analysis", "Index", "Dashboard"]

# ── row-keyed diff config ──────────────────────────────────────────────────
KEY_COL = 0          # 0-based column holding the stable row id (player name), AFTER any
                     #   leading positional-index column is dropped
HEADER_ROW = 0       # 0-based row index of the column headers (set to None if no header row)
IGNORE_COLS = []     # header names NOT to track (volatile/formula cols), e.g. ["AGE", "TODAY"]
MAX_FIELDS_PER_ROW = 6   # cap fields listed for one changed row (keeps the line readable)



def sheet_key(url):
    m = re.search(r"/spreadsheets/d/([^/]+)", url)
    return m.group(1) if m else re.sub(r"[^A-Za-z0-9]+", "_", url)[:40]


def xlsx_url(sid):
    return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"


def fetch_bytes(url):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (nba-sheets-tracker)"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read()


IDX_RE = re.compile(r"^\d+(?:\.0+)?$")


def _leading_index(raw):
    """True if column 0 is a sequential row counter (4,5,6,... or 4.0,5.0,...)."""
    nums = []
    for cells in raw:
        if cells and IDX_RE.match(cells[0]):
            try:
                nums.append(float(cells[0]))
            except ValueError:
                pass
    if len(nums) < 3:
        return False
    inc = sum(1 for i in range(1, len(nums)) if abs(nums[i] - nums[i - 1] - 1) < 1e-6)
    return inc >= (len(nums) - 1) * 0.6


def read_workbook(data):
    """Return {tab: {"h": [headers], "r": {player_key: [cells]}}} for every tab.
    Row-keyed by player name so a cell edit maps to ONE row and inserting/deleting a row
    no longer cascades. Trailing empties trimmed; a leading index column dropped."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = {}
    for ws in wb.worksheets:
        if ws.title in SKIP_TABS:
            continue
        raw = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if any(cells):
                raw.append(cells)
        # If column 0 is a positional row counter, drop it — otherwise deleting one row
        # renumbers everything below it and every row shows up as a change.
        if _leading_index(raw):
            raw = [cells[1:] for cells in raw]
        if not raw:
            out[ws.title] = {"h": [], "r": {}}
            continue
        if HEADER_ROW is not None and 0 <= HEADER_ROW < len(raw):
            headers, body = raw[HEADER_ROW], raw[HEADER_ROW + 1:]
        else:
            headers, body = [], raw
        rows = {}
        for cells in body:
            key = cells[KEY_COL] if KEY_COL < len(cells) else ""
            if not key:
                continue
            k, n = key, 2
            while k in rows:                 # de-dupe identical names so neither is lost
                k, n = f"{key} #{n}", n + 1
            rows[k] = cells
        out[ws.title] = {"h": headers, "r": rows}
    wb.close()
    return out


def diff_tab(old_tab, new_tab):
    """Concise row-keyed diff of one tab -> [(type, text), ...] with ONE entry per added /
    removed / changed player. Changed rows list only the cells that actually moved."""
    out = []
    old_r = (old_tab or {}).get("r", {})
    new_r = (new_tab or {}).get("r", {})
    headers = (new_tab or {}).get("h", [])
    for key in new_r:
        if key not in old_r:
            out.append(("ADDED", key))
    for key in old_r:
        if key not in new_r:
            out.append(("REMOVED", key))
    for key, ncells in new_r.items():
        ocells = old_r.get(key)
        if ocells is None:
            continue
        changed = []
        for i in range(max(len(ocells), len(ncells))):
            if i == KEY_COL:
                continue
            ov = ocells[i] if i < len(ocells) else ""
            nv = ncells[i] if i < len(ncells) else ""
            if ov == nv:
                continue
            label = headers[i] if i < len(headers) and headers[i] else f"col{i + 1}"
            if label in IGNORE_COLS:
                continue
            changed.append(f"{label}: {ov or '-'} -> {nv or '-'}")
        if changed:
            shown = changed[:MAX_FIELDS_PER_ROW]
            if len(changed) > MAX_FIELDS_PER_ROW:
                shown.append(f"(+{len(changed) - MAX_FIELDS_PER_ROW} more)")
            out.append(("CHANGED", f"{key}  -  " + ", ".join(shown)))
    return out


def load_prev(key):
    p = os.path.join(SNAP, key + ".json")
    if os.path.exists(p):
        try:
            return json.load(open(p, encoding="utf-8"))
        except Exception:
            return None
    return None


def save_snap(key, tabs):
    json.dump({"tabs": tabs}, open(os.path.join(SNAP, key + ".json"), "w", encoding="utf-8"),
              ensure_ascii=False)


def append_log(rows):
    new = not os.path.exists(LOG)
    with open(LOG, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if new:
            w.writerow(["detected_at", "sheet", "type", "text"])
        for r in rows:
            w.writerow(r)


def _safe_key(name):
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name)[:80]


def poll_once(silent=False, dry=False):
    try:
        import openpyxl  # noqa: F401
    except Exception:
        print("openpyxl is required for sheet tracking. Install it with:  pip install openpyxl")
        return

    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_changes = []
    per_tab = []

    for friendly, url in SPREADSHEETS:
        sid = sheet_key(url)
        try:
            tabs = read_workbook(fetch_bytes(xlsx_url(sid)))
        except Exception as e:
            print(f"fetch/parse failed [{friendly or sid}]: {e}")
            continue

        snap_key = _safe_key(sid)
        prev = (load_prev(snap_key) or {}).get("tabs", {})
        # Old snapshots stored each tab as a list of row-strings; the new format is a keyed
        # dict. If we see the old shape, re-baseline instead of flooding with fake diffs.
        if prev and any(not isinstance(v, dict) for v in prev.values()):
            prev = {}
        save_snap(snap_key, tabs)

        if not prev:
            rowcount = sum(len(t.get("r", {})) for t in tabs.values())
            print(f"baseline saved: {friendly or sid} \u2014 {len(tabs)} tab(s), {rowcount} rows")
            continue

        for tab, tabdata in tabs.items():
            label = f"{friendly} \u00b7 {tab}" if friendly else tab
            changes = diff_tab(prev.get(tab, {}), tabdata)
            if dry and (changes or tab not in prev):
                tag = " (NEW TAB)" if tab not in prev else ""
                print(f"[{label}] {len(tabdata.get('r', {}))} rows, {len(changes)} change(s){tag}")
            if changes:
                per_tab.append((label, len(changes)))
                for typ, text in changes:
                    all_changes.append((now, label, typ, text))
        for tab in prev:
            if tab not in tabs:
                label = f"{friendly} \u00b7 {tab}" if friendly else tab
                print(f"[{label}] tab removed")

    if all_changes and not dry:
        append_log(all_changes)

    if all_changes:
        summary = ", ".join(f"{name} ({n})" for name, n in per_tab)
        print(f"{len(all_changes)} change(s) across {len(per_tab)} tab(s): {summary}")
        if HAVE_TOAST and not dry and not silent:
            try:
                head = per_tab[0][0] if len(per_tab) == 1 else f"{len(per_tab)} tabs"
                t = Notification(app_id="NBA Sheets Tracker",
                                 title=f"Sheet update: {len(all_changes)} change(s)",
                                 msg=f"{head} \u2014 {summary}"[:250])
                t.set_audio(audio.Default, loop=False)
                t.show()
            except Exception:
                pass
    else:
        print("no changes" if not dry else "(test) no changes vs last snapshot")


def send_test():
    """Sample notification (used by nba_tracker_all.py --test)."""
    if HAVE_TOAST:
        try:
            t = Notification(app_id="NBA Sheets Tracker", title="NBA Sheets Tracker",
                             msg="Sheet tracking is wired up and working.")
            t.set_audio(audio.Default, loop=False)
            t.show()
            print("sent sample toast")
        except Exception as e:
            print(f"toast failed: {e}")
    else:
        print("winotify not installed (pip install winotify) - no toast sent")


def main():
    poll_once(silent="--silent" in sys.argv, dry="--test" in sys.argv)


if __name__ == "__main__":
    main()